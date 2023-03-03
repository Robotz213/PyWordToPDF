import os
import openpyxl
from subprocess import Popen
import threading
import comtypes.client
from pathlib import *
from checkword import checkdir
from termcolor import colored
import sys

nomexls = input(colored("Informe a Planilha: ", "yellow"))


class ConvertWordPDF:

    def initbot(self):
            self.fontdir = Path(__file__).parent.resolve()
            p = os.path.join(self.fontdir, nomexls)


            wrkbk_input = openpyxl.load_workbook(filename=p)
            sheet_input = wrkbk_input.active
                    

            for i in range(1, sheet_input.max_row+1):
                    cell_obj = sheet_input.cell(row=i, column=1)
                    if cell_obj.value is not None and cell_obj.value != '':
                        docx_name = sheet_input.cell(row=i, column=1).value


                        wordir = checkdir().find_all(
                             exefile='winword.exe', 
                             pathfile=r'c:/Program Files')
                        if wordir is True:
                            try:
                                self.convertpdf(namefile=docx_name)
                            except Exception as e:
                                print(e)
                        else:
                             self.pdfinsoffice(namefile=docx_name)
        
                        if i == sheet_input.max_row:
                            print('fim do processo')
                            os.system("PAUSE")

    def convertpdf(self, namefile):

        wdFormatPDF = 17
        currentdir = Path(__file__).parent.resolve()
        inputfile = os.path.join(currentdir, '{infile}'.format(infile=namefile))
        print(inputfile) 
        removedocx = inputfile.replace(".docx",".pdf")
        outpatch = removedocx
        
        
        # Cria instancia de um objeto COM para manipular Documentos Word
        word = comtypes.client.CreateObject('Word.Application')

        # Carrega Arquivo de entrada (.doc)
        doc = word.Documents.Open(inputfile)
        print(colored('Convertendo arquivo "{a}"...'.format(a=namefile), 'blue'))

        try:
            
            # Salva arquivo de saida em formato .pdf
            doc.SaveAs(outpatch, FileFormat=wdFormatPDF)

            # Fecha arquivo de Entrada
            doc.Close()

            # Finaliza instancia do Objeto COM criado
            word.Quit()
            print(colored('Documento Convertido com Sucesso!', 'green'))

        except:
            print(colored('Não foi possível converter o arquivo "{b}"'.formar(b=namefile), 'red'))  

    def pdfinsoffice(self, namefile):

        path = 'c:\\Program Files\\'
        name = 'soffice.exe'

        if sys.platform == 'linux':
            currentdir = Path(__file__).parent.resolve()
            arquivo_de_entrada = os.path.join(currentdir, '{infile}'.format(infile=namefile))
            outfile = os.path.join(currentdir)
            
            print(colored('Convertendo arquivo "{a}"...'.format(a=namefile), 'blue'))
            
            try:
                p = Popen(['soffice', '--headless', '--convert-to', 'pdf', '--outdir', outfile , arquivo_de_entrada])
                p.communicate()
                print(colored('Documento Convertido com Sucesso!', 'green'))
            except:
                print(colored('Não foi possível converter o arquivo "{b}"'.formar(b=namefile), 'red'))
        else:
            LIBRE_OFFICE = ''
            for root, dirs, files in os.walk(path):
                if name in files:
                    LIBRE_OFFICE = os.path.join(root, name)

            currentdir = Path(__file__).parent.resolve()
            arquivo_de_entrada = os.path.join(currentdir, '{infile}'.format(infile=namefile))
            outfile = os.path.join(currentdir)
            
            print(colored('Convertendo arquivo "{a}"...'.format(a=namefile), 'blue'))
            
            try:
                p = Popen([LIBRE_OFFICE, '--headless', '--convert-to', 'pdf', '--outdir', outfile , arquivo_de_entrada])
                p.communicate()
                print(colored('Documento Convertido com Sucesso!', 'green'))
            except:
                print(colored('Não foi possível converter o arquivo "{b}"'.formar(b=namefile), 'red'))
             
            

         
start = ConvertWordPDF()
start.initbot()

